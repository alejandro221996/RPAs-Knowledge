"""
Ejemplo 4: Procesamiento Automatizado de Archivos
=================================================

Este ejemplo demuestra cómo automatizar el procesamiento de archivos Excel
usando pandas y openpyxl. Es útil para:
- Procesamiento automático de reportes
- Generación de archivos Excel
- Organización y limpieza de datos
- Automatización de tareas administrativas
"""

import pandas as pd
import os
import shutil
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class FileProcessingRPA:
    def __init__(self):
        """Inicializa el bot de procesamiento de archivos"""
        self.processed_files = []
        self.reports_generated = []
        
    def create_sample_data(self):
        """Crea datos de ejemplo para demostrar el procesamiento"""
        print("📊 Creando datos de ejemplo...")
        
        # Crear directorio de trabajo si no existe
        if not os.path.exists('datos_ejemplo'):
            os.makedirs('datos_ejemplo')
            
        # 1. Datos de ventas
        sales_data = []
        products = ['Laptop', 'Mouse', 'Teclado', 'Monitor', 'Auriculares']
        regions = ['Norte', 'Sur', 'Este', 'Oeste', 'Centro']
        
        for i in range(100):
            sale = {
                'ID_Venta': f'V{i+1:03d}',
                'Producto': random.choice(products),
                'Cantidad': random.randint(1, 10),
                'Precio_Unitario': round(random.uniform(50, 2000), 2),
                'Region': random.choice(regions),
                'Fecha': (datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d'),
                'Vendedor': f'Vendedor{random.randint(1, 10)}'
            }
            sale['Total'] = sale['Cantidad'] * sale['Precio_Unitario']
            sales_data.append(sale)
            
        # Guardar datos de ventas
        df_sales = pd.DataFrame(sales_data)
        df_sales.to_excel('datos_ejemplo/ventas_2024.xlsx', index=False)
        print("   ✅ Archivo de ventas creado: datos_ejemplo/ventas_2024.xlsx")
        
        # 2. Datos de inventario
        inventory_data = []
        for i, product in enumerate(products):
            inventory = {
                'ID_Producto': f'P{i+1:03d}',
                'Producto': product,
                'Stock_Actual': random.randint(10, 100),
                'Stock_Minimo': 20,
                'Precio_Costo': round(random.uniform(30, 1500), 2),
                'Proveedor': f'Proveedor{random.randint(1, 5)}',
                'Categoria': random.choice(['Electrónicos', 'Accesorios', 'Periféricos'])
            }
            inventory_data.append(inventory)
            
        # Guardar datos de inventario
        df_inventory = pd.DataFrame(inventory_data)
        df_inventory.to_excel('datos_ejemplo/inventario_2024.xlsx', index=False)
        print("   ✅ Archivo de inventario creado: datos_ejemplo/inventario_2024.xlsx")
        
        # 3. Datos de empleados
        employees_data = []
        departments = ['Ventas', 'Marketing', 'IT', 'Recursos Humanos', 'Finanzas']
        
        for i in range(20):
            employee = {
                'ID_Empleado': f'E{i+1:03d}',
                'Nombre': f'Empleado{i+1}',
                'Apellido': f'Apellido{i+1}',
                'Departamento': random.choice(departments),
                'Salario': round(random.uniform(30000, 80000), 2),
                'Fecha_Contratacion': (datetime.now() - timedelta(days=random.randint(100, 1000))).strftime('%Y-%m-%d'),
                'Email': f'empleado{i+1}@empresa.com'
            }
            employees_data.append(employee)
            
        # Guardar datos de empleados
        df_employees = pd.DataFrame(employees_data)
        df_employees.to_excel('datos_ejemplo/empleados_2024.xlsx', index=False)
        print("   ✅ Archivo de empleados creado: datos_ejemplo/empleados_2024.xlsx")
        
        print(f"📁 Total de archivos de ejemplo creados: 3")
        
    def process_sales_data(self):
        """Procesa los datos de ventas y genera reportes"""
        print("\n📈 Procesando datos de ventas...")
        
        try:
            # Leer archivo de ventas
            df_sales = pd.read_excel('datos_ejemplo/ventas_2024.xlsx')
            
            # Análisis de ventas por región
            sales_by_region = df_sales.groupby('Region').agg({
                'Total': ['sum', 'mean', 'count'],
                'Cantidad': 'sum'
            }).round(2)
            
            # Análisis de ventas por producto
            sales_by_product = df_sales.groupby('Producto').agg({
                'Total': ['sum', 'mean'],
                'Cantidad': 'sum'
            }).round(2)
            
            # Análisis de ventas por fecha
            df_sales['Fecha'] = pd.to_datetime(df_sales['Fecha'])
            sales_by_date = df_sales.groupby(df_sales['Fecha'].dt.to_period('M')).agg({
                'Total': 'sum',
                'ID_Venta': 'count'
            }).round(2)
            
            # Crear reporte consolidado
            report_data = {
                'Resumen_General': {
                    'Total_Ventas': len(df_sales),
                    'Ingresos_Totales': df_sales['Total'].sum(),
                    'Promedio_Venta': df_sales['Total'].mean(),
                    'Producto_Mas_Vendido': df_sales.groupby('Producto')['Cantidad'].sum().idxmax(),
                    'Region_Mas_Activa': df_sales.groupby('Region')['Total'].sum().idxmax()
                },
                'Ventas_por_Region': sales_by_region,
                'Ventas_por_Producto': sales_by_product,
                'Ventas_por_Mes': sales_by_date
            }
            
            # Guardar reporte
            self.save_sales_report(report_data)
            
            print("   ✅ Procesamiento de ventas completado")
            return report_data
            
        except Exception as e:
            print(f"   ❌ Error procesando ventas: {e}")
            return None
            
    def process_inventory_data(self):
        """Procesa los datos de inventario y genera alertas"""
        print("\n📦 Procesando datos de inventario...")
        
        try:
            # Leer archivo de inventario
            df_inventory = pd.read_excel('datos_ejemplo/inventario_2024.xlsx')
            
            # Calcular valor total del inventario
            df_inventory['Valor_Total'] = df_inventory['Stock_Actual'] * df_inventory['Precio_Costo']
            
            # Identificar productos con stock bajo
            low_stock = df_inventory[df_inventory['Stock_Actual'] <= df_inventory['Stock_Minimo']]
            
            # Análisis por categoría
            inventory_by_category = df_inventory.groupby('Categoria').agg({
                'Stock_Actual': 'sum',
                'Valor_Total': 'sum',
                'ID_Producto': 'count'
            }).round(2)
            
            # Análisis por proveedor
            inventory_by_supplier = df_inventory.groupby('Proveedor').agg({
                'Stock_Actual': 'sum',
                'Valor_Total': 'sum'
            }).round(2)
            
            # Crear reporte de inventario
            inventory_report = {
                'Resumen_Inventario': {
                    'Total_Productos': len(df_inventory),
                    'Valor_Total_Inventario': df_inventory['Valor_Total'].sum(),
                    'Productos_Stock_Bajo': len(low_stock),
                    'Categoria_Mayor_Valor': df_inventory.groupby('Categoria')['Valor_Total'].sum().idxmax()
                },
                'Productos_Stock_Bajo': low_stock,
                'Inventario_por_Categoria': inventory_by_category,
                'Inventario_por_Proveedor': inventory_by_supplier
            }
            
            # Guardar reporte
            self.save_inventory_report(inventory_report)
            
            print("   ✅ Procesamiento de inventario completado")
            return inventory_report
            
        except Exception as e:
            print(f"   ❌ Error procesando inventario: {e}")
            return None
            
    def save_sales_report(self, report_data):
        """Guarda el reporte de ventas con formato profesional"""
        print("   💾 Guardando reporte de ventas...")
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen General
        ws1 = wb.active
        ws1.title = "Resumen General"
        
        # Título
        ws1['A1'] = "REPORTE DE VENTAS - 2024"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        # Datos del resumen
        row = 3
        for key, value in report_data['Resumen_General'].items():
            ws1[f'A{row}'] = key.replace('_', ' ')
            ws1[f'B{row}'] = value
            row += 1
            
        # Hoja 2: Ventas por Región
        ws2 = wb.create_sheet("Ventas por Región")
        for r in dataframe_to_rows(report_data['Ventas_por_Region'], index=True, header=True):
            ws2.append(r)
            
        # Hoja 3: Ventas por Producto
        ws3 = wb.create_sheet("Ventas por Producto")
        for r in dataframe_to_rows(report_data['Ventas_por_Producto'], index=True, header=True):
            ws3.append(r)
            
        # Guardar archivo
        filename = f"reportes/reporte_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs('reportes', exist_ok=True)
        wb.save(filename)
        
        self.reports_generated.append(filename)
        print(f"   ✅ Reporte guardado: {filename}")
        
    def save_inventory_report(self, report_data):
        """Guarda el reporte de inventario con alertas"""
        print("   💾 Guardando reporte de inventario...")
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen Inventario"
        
        # Título
        ws1['A1'] = "REPORTE DE INVENTARIO - 2024"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        # Datos del resumen
        row = 3
        for key, value in report_data['Resumen_Inventario'].items():
            ws1[f'A{row}'] = key.replace('_', ' ')
            ws1[f'B{row}'] = value
            row += 1
            
        # Hoja 2: Alertas de Stock Bajo
        ws2 = wb.create_sheet("Alertas Stock Bajo")
        if not report_data['Productos_Stock_Bajo'].empty:
            for r in dataframe_to_rows(report_data['Productos_Stock_Bajo'], index=False, header=True):
                ws2.append(r)
                
            # Resaltar filas con stock bajo
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else:
            ws2['A1'] = "No hay productos con stock bajo"
            
        # Hoja 3: Inventario por Categoría
        ws3 = wb.create_sheet("Inventario por Categoría")
        for r in dataframe_to_rows(report_data['Inventario_por_Categoria'], index=True, header=True):
            ws3.append(r)
            
        # Guardar archivo
        filename = f"reportes/reporte_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs('reportes', exist_ok=True)
        wb.save(filename)
        
        self.reports_generated.append(filename)
        print(f"   ✅ Reporte guardado: {filename}")
        
    def organize_files(self):
        """Organiza los archivos procesados en carpetas"""
        print("\n📁 Organizando archivos...")
        
        # Crear estructura de carpetas
        folders = ['datos_originales', 'datos_procesados', 'reportes', 'archivos_temporales']
        
        for folder in folders:
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"   📂 Carpeta creada: {folder}")
                
        # Mover archivos originales
        source_files = [
            'datos_ejemplo/ventas_2024.xlsx',
            'datos_ejemplo/inventario_2024.xlsx',
            'datos_ejemplo/empleados_2024.xlsx'
        ]
        
        for file in source_files:
            if os.path.exists(file):
                shutil.move(file, 'datos_originales/')
                print(f"   📄 Movido: {file} → datos_originales/")
                
        # Crear archivo de log de procesamiento
        log_data = {
            'Archivo': [],
            'Tipo_Procesamiento': [],
            'Fecha_Procesamiento': [],
            'Estado': []
        }
        
        for file in source_files:
            filename = os.path.basename(file)
            if 'ventas' in filename:
                log_data['Archivo'].append(filename)
                log_data['Tipo_Procesamiento'].append('Análisis de Ventas')
                log_data['Fecha_Procesamiento'].append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                log_data['Estado'].append('Completado')
            elif 'inventario' in filename:
                log_data['Archivo'].append(filename)
                log_data['Tipo_Procesamiento'].append('Análisis de Inventario')
                log_data['Fecha_Procesamiento'].append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                log_data['Estado'].append('Completado')
                
        # Guardar log
        df_log = pd.DataFrame(log_data)
        df_log.to_excel('archivos_temporales/log_procesamiento.xlsx', index=False)
        print("   📋 Log de procesamiento creado")
        
    def cleanup_temp_files(self):
        """Limpia archivos temporales"""
        print("\n🧹 Limpiando archivos temporales...")
        
        temp_files = [
            'archivos_temporales/log_procesamiento.xlsx'
        ]
        
        for file in temp_files:
            if os.path.exists(file):
                os.remove(file)
                print(f"   🗑️ Eliminado: {file}")
                
    def run_file_processing(self):
        """Ejecuta el proceso completo de procesamiento de archivos"""
        print("🚀 Iniciando proceso de Procesamiento de Archivos RPA")
        print("=" * 60)
        
        try:
            # 1. Crear datos de ejemplo
            self.create_sample_data()
            
            # 2. Procesar datos de ventas
            sales_report = self.process_sales_data()
            
            # 3. Procesar datos de inventario
            inventory_report = self.process_inventory_data()
            
            # 4. Organizar archivos
            self.organize_files()
            
            # 5. Limpiar archivos temporales
            self.cleanup_temp_files()
            
            print("\n🎉 Proceso de procesamiento de archivos completado!")
            print(f"📊 Reportes generados: {len(self.reports_generated)}")
            print(f"📁 Archivos procesados: {len(self.processed_files)}")
            
            # Mostrar resumen de reportes
            if self.reports_generated:
                print("\n📋 Reportes generados:")
                for report in self.reports_generated:
                    print(f"   📄 {report}")
                    
        except Exception as e:
            print(f"❌ Error en el proceso: {e}")

def main():
    """Función principal para ejecutar el ejemplo"""
    print("🤖 RPA File Processing - Ejemplo de Procesamiento de Archivos")
    print("=" * 60)
    
    # Crear instancia del bot
    file_bot = FileProcessingRPA()
    
    # Ejecutar proceso completo
    file_bot.run_file_processing()

if __name__ == "__main__":
    main() 